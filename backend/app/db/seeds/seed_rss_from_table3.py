"""Seed: 新闻源.xlsx → SourceRegistry (source_type='rss')。

3 个 sheet × 14-16 行 RSS Feed。每行 → 一条 SourceRegistry。
- sheet 名 → direction_tags（AI人工智能 / 财经 / 国际政治）
- 权重列 → weight
"""

import asyncio
import logging
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import select

from app.db.seeds import TABLE3_PATH
from app.db.session import AsyncSessionLocal
from app.models.source_registry import SourceRegistry, SOURCE_TYPE_RSS

logger = logging.getLogger(__name__)


SHEET_TO_DIRECTION = {
    "AI人工智能": "AI",
    "财经": "财经",
    "国际政治": "国际政治",
}


def _slug_platform(name: str) -> str:
    """生成 platform：保留中文，去除特殊字符，便于人读。"""
    # 简化：取前 40 字，把空格/标点换成 _
    import re
    s = re.sub(r"[\s/\\()（）【】\.,，。、]+", "_", name).strip("_")[:60]
    return f"rss_{s}"


async def _upsert(db, *, name: str, url: str, weight: int, direction: str) -> str:
    platform = _slug_platform(name)
    existing = (await db.execute(
        select(SourceRegistry).where(SourceRegistry.platform == platform)
    )).scalar_one_or_none()

    tags = [direction]
    if existing:
        existing.name = name
        existing.url = url
        existing.source_type = SOURCE_TYPE_RSS
        existing.weight = weight
        # 合并 direction_tags
        merged = list({*(existing.direction_tags or []), direction})
        existing.direction_tags = merged
        existing.enabled = True
        return "updated"

    db.add(SourceRegistry(
        name=name,
        platform=platform,
        source_type=SOURCE_TYPE_RSS,
        url=url,
        tier=None,
        direction_tags=tags,
        weight=weight,
        requires_auth=False,
        auth_status="ok",
        fetch_strategy="cron",
        fetch_config={},
        enabled=True,
        description=f"RSS 源（{direction}）",
    ))
    return "created"


def _to_int(value, default: int = 5) -> int:
    if value is None:
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


async def run(db) -> dict:
    if not Path(TABLE3_PATH).exists():
        raise FileNotFoundError(f"Seed file missing: {TABLE3_PATH}")

    wb = openpyxl.load_workbook(TABLE3_PATH, data_only=True)
    stats = {"created": 0, "updated": 0, "skipped": 0}

    for sheet_name in wb.sheetnames:
        direction = SHEET_TO_DIRECTION.get(sheet_name, sheet_name)
        ws = wb[sheet_name]
        # 表头在 R2: 来源名称 / RSS URL / 权重；数据从 R3 开始
        for row in ws.iter_rows(min_row=3, values_only=True):
            cols = (row + (None, None, None))[:3]
            name, url, weight = cols
            if not name or not url:
                stats["skipped"] += 1
                continue
            outcome = await _upsert(
                db,
                name=str(name).strip(),
                url=str(url).strip(),
                weight=_to_int(weight),
                direction=direction,
            )
            stats[outcome] += 1

    await db.commit()
    logger.info(f"seed_rss_from_table3: {stats}")
    return stats


async def _main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    async with AsyncSessionLocal() as db:
        result = await run(db)
    print(f"Done: {result}")


if __name__ == "__main__":
    asyncio.run(_main())
